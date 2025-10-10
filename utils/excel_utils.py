"""
Utilidades para generación de archivos Excel de facturas.
"""

from typing import Dict, Any
from datetime import datetime
from io import BytesIO
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("excel_utils")


def generar_excel_factura(datos: Dict[str, Any], filename: str = None) -> BytesIO:
    """
    Genera un archivo Excel con los datos de una factura.
    
    Args:
        datos: Diccionario con datos de la factura (formato del frontend)
        filename: Nombre del archivo PDF original (opcional)
    
    Returns:
        BytesIO: Buffer con el archivo Excel generado
    """
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Factura"
        
        # === ESTILOS ===
        titulo_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        titulo_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        label_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=10)
        
        total_font = Font(name='Arial', size=11, bold=True)
        total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # === TÍTULO PRINCIPAL ===
        row = 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "FACTURA - DATOS EXTRAÍDOS"
        cell.font = titulo_font
        cell.fill = titulo_fill
        cell.alignment = center_alignment
        ws.row_dimensions[row].height = 25
        
        row += 1
        
        # === INFORMACIÓN DEL DOCUMENTO ===
        if filename:
            ws[f'A{row}'] = "Archivo origen:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = filename
            ws[f'B{row}'].font = normal_font
            row += 1
        
        ws[f'A{row}'] = "Fecha de generación:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        ws[f'B{row}'].font = normal_font
        row += 2
        
        # === DATOS DEL CLIENTE ===
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "DATOS DEL CLIENTE"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        row += 1
        
        # Código Cliente
        ws[f'A{row}'] = "Código Cliente:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = datos.get('codigo_cliente') or 'N/A'
        ws[f'B{row}'].font = normal_font
        row += 1
        
        # Razón Social
        ws[f'A{row}'] = "Razón Social:"
        ws[f'A{row}'].font = label_font
        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = datos.get('razon_social_cliente') or 'N/A'
        ws[f'B{row}'].font = normal_font
        row += 1
        
        # Dirección
        ws[f'A{row}'] = "Dirección:"
        ws[f'A{row}'].font = label_font
        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = datos.get('direccion_cliente') or 'N/A'
        ws[f'B{row}'].font = normal_font
        row += 1
        
        # Distrito
        ws[f'A{row}'] = "Distrito:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = datos.get('distrito') or 'N/A'
        ws[f'B{row}'].font = normal_font
        row += 2
        
        # === ITEMS (TABLA) ===
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "DETALLE DE ITEMS"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        row += 1
        
        # Encabezados de tabla
        headers = ['#', 'Descripción', 'Cantidad', 'Precio Unitario', 'Subtotal']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = center_alignment
            cell.border = border
        
        row += 1
        
        # Items
        items = datos.get('items', [])
        if items:
            for idx, item in enumerate(items, 1):
                ws.cell(row=row, column=1, value=idx).alignment = center_alignment
                ws.cell(row=row, column=1).border = border
                
                ws.cell(row=row, column=2, value=item.get('descripcion', 'N/A')).alignment = left_alignment
                ws.cell(row=row, column=2).border = border
                
                ws.cell(row=row, column=3, value=item.get('cantidad', 0)).alignment = center_alignment
                ws.cell(row=row, column=3).border = border
                
                precio = item.get('precio_unitario', 0)
                ws.cell(row=row, column=4, value=precio).alignment = right_alignment
                ws.cell(row=row, column=4).number_format = '#,##0.00'
                ws.cell(row=row, column=4).border = border
                
                subtotal = item.get('subtotal', 0)
                ws.cell(row=row, column=5, value=subtotal).alignment = right_alignment
                ws.cell(row=row, column=5).number_format = '#,##0.00'
                ws.cell(row=row, column=5).border = border
                
                row += 1
        else:
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = "Sin items registrados"
            ws[f'A{row}'].alignment = center_alignment
            ws[f'A{row}'].font = Font(italic=True)
            row += 1
        
        row += 1
        
        # === TOTALES ===
        moneda = datos.get('moneda', 'PEN')
        
        # Subtotal
        ws[f'D{row}'] = "Subtotal:"
        ws[f'D{row}'].font = label_font
        ws[f'D{row}'].alignment = right_alignment
        ws[f'E{row}'] = datos.get('subtotal', 0)
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].number_format = f'"{moneda}" #,##0.00'
        ws[f'E{row}'].alignment = right_alignment
        row += 1
        
        # IGV
        ws[f'D{row}'] = "IGV (18%):"
        ws[f'D{row}'].font = label_font
        ws[f'D{row}'].alignment = right_alignment
        ws[f'E{row}'] = datos.get('igv', 0)
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].number_format = f'"{moneda}" #,##0.00'
        ws[f'E{row}'].alignment = right_alignment
        row += 1
        
        # Total
        ws[f'D{row}'] = "TOTAL:"
        ws[f'D{row}'].font = total_font
        ws[f'D{row}'].fill = total_fill
        ws[f'D{row}'].alignment = right_alignment
        ws[f'D{row}'].border = border
        
        ws[f'E{row}'] = datos.get('total', 0)
        ws[f'E{row}'].font = total_font
        ws[f'E{row}'].fill = total_fill
        ws[f'E{row}'].number_format = f'"{moneda}" #,##0.00'
        ws[f'E{row}'].alignment = right_alignment
        ws[f'E{row}'].border = border
        row += 2
        
        # === DETRACCIÓN (si existe) ===
        detraccion = datos.get('detraccion')
        if detraccion and (detraccion.get('porcentaje', 0) > 0 or detraccion.get('monto', 0) > 0):
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = "DETRACCIÓN"
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            row += 1
            
            # Porcentaje
            ws[f'D{row}'] = "Porcentaje:"
            ws[f'D{row}'].font = label_font
            ws[f'D{row}'].alignment = right_alignment
            ws[f'E{row}'] = f"{detraccion.get('porcentaje', 0)}%"
            ws[f'E{row}'].font = normal_font
            ws[f'E{row}'].alignment = right_alignment
            row += 1
            
            # Monto
            ws[f'D{row}'] = "Monto Detracción:"
            ws[f'D{row}'].font = label_font
            ws[f'D{row}'].alignment = right_alignment
            ws[f'E{row}'] = detraccion.get('monto', 0)
            ws[f'E{row}'].font = normal_font
            ws[f'E{row}'].number_format = f'"{moneda}" #,##0.00'
            ws[f'E{row}'].alignment = right_alignment
            row += 1
        
        row += 2
        
        # === INFORMACIÓN ADICIONAL ===
        ws[f'A{row}'] = "Forma de Pago:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = datos.get('forma_pago') or 'N/A'
        ws[f'B{row}'].font = normal_font
        row += 1
        
        ws[f'A{row}'] = "Moneda:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = moneda
        ws[f'B{row}'].font = normal_font
        
        # === AJUSTAR ANCHOS DE COLUMNA ===
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # === GUARDAR EN BUFFER ===
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Excel generado exitosamente para factura")
        return buffer
    
    except Exception as e:
        logger.error(f"Error generando Excel: {e}")
        raise
